from rest_framework import generics, status, viewsets, filters
from django.http import JsonResponse
from rest_framework.response import Response
from rest_framework.decorators import (
    api_view,
    permission_classes,
    action,
    authentication_classes,
)
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.authentication import TokenAuthentication
from django.contrib.auth import authenticate, login
from rest_framework.authtoken.models import Token
from .models import (
    Lead,
    Student,
    LeadRemark,
    WalkIn,
    StudentChangeLog,
    UserProfile,
    ActivityLog,
    Invoice,
    Receipt,
    LoginAttempt,
    Company,
    StaffUser,
    CustomAdmissionField,
    StudentCustomFieldValue,
)
from .utils import get_client_ip
from .serializers import (
    LeadSerializer,
    StudentSerializer,
    LeadRemarkSerializer,
    WalkInSerializer,
    UserSerializer,
    CompanySerializer,
    StaffUserSerializer,
    CustomAdmissionFieldSerializer,
    StudentCustomFieldValueSerializer,
)
from .activity_logger import log_activity
from django.contrib.auth.models import User
from django.db.models import Count, Sum
from datetime import datetime, timedelta
from django.utils import timezone
import calendar
import os
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from django.http import HttpResponse
import io

from django.views.decorators.csrf import csrf_exempt


# Login View
@csrf_exempt
@api_view(["POST"])
@permission_classes([AllowAny])
@authentication_classes([])
def login_view(request):
    ip = get_client_ip(request)
    username = request.data.get("username")
    password = request.data.get("password")

    # Check for blocking condition
    failed_attempts = LoginAttempt.objects.filter(
        username=username,
        ip_address=ip,
        status="FAILED",
        timestamp__gte=timezone.now() - timedelta(minutes=15),
    ).count()

    if failed_attempts >= 5:
        return Response(
            {"error": "Too many failed attempts. Try again after 15 minutes."},
            status=status.HTTP_403_FORBIDDEN,
        )

    import jwt
    from django.conf import settings

    def generate_jwt(user_obj, role):
        payload = {
            "user_id": user_obj.id,
            "username": user_obj.username,
            "email": getattr(user_obj, "email", ""),
            "role": role,
            "exp": datetime.utcnow() + timedelta(days=7),
        }
        return jwt.encode(payload, settings.SECRET_KEY, algorithm="HS256")

    print(f"DEBUG: Login attempt for username='{username}' from IP={ip}")
    
    # Step 1: Authenticate the user
    user = authenticate(username=username, password=password)
    
    if user is not None:
        LoginAttempt.objects.create(username=username, ip_address=ip, status="SUCCESS")

        # Get role from the StaffUser or UserProfile
        from lead_enrollment.models import StaffUser, UserProfile

        staff = StaffUser.objects.filter(user_id=user.id).first()
        if staff:
            role = staff.role
        else:
            try:
                profile = UserProfile.objects.get(user=user)
                role = profile.role
            except UserProfile.DoesNotExist:
                role = "STAFF"  # fallback

        token_jwt = generate_jwt(user, role)

        return Response(
            {
                "token": token_jwt,
                "email": user.email,
                "role": role,
                "user": {
                    "id": user.id,
                    "username": user.username,
                    "email": user.email,
                    "role": role,
                    "redirect_url": "/dashboard",
                },
            }
        )
    else:
        LoginAttempt.objects.create(username=username, ip_address=ip, status="FAILED")
        return Response(
            {"error": "Invalid Credentials"}, status=status.HTTP_400_BAD_REQUEST
        )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def profile_view(request):
    from lead_enrollment.models import StaffUser, Company

    company_name = "System"
    try:
        company_id = user.userprofile.company_id
        if company_id > 0:
            company = Company.objects.get(id=company_id)
            company_name = company.company_name
    except Exception:
        company_id = 0

    staff = StaffUser.objects.filter(user_id=user.id).first()
    if staff:
        role = staff.role
    else:
        role = user.userprofile.role if hasattr(user, "userprofile") else "STAFF"

    return Response(
        {
            "username": user.username,
            "email": user.email,
            "role": role,
            "company_id": company_id,
            "company_name": company_name,
            "last_login": user.last_login,
        }
    )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def change_password(request):
    user = request.user
    old_password = request.data.get("old_password")
    new_password = request.data.get("new_password")

    if not user.check_password(old_password):
        return Response(
            {"error": "Incorrect old password"}, status=status.HTTP_400_BAD_REQUEST
        )

    user.set_password(new_password)
    user.save()
    return Response({"message": "Password changed successfully"})


class CompanyViewSet(viewsets.ModelViewSet):
    queryset = Company.objects.all()
    serializer_class = CompanySerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        if self.request.user.userprofile.role == "SUPERADMIN":
            return Company.objects.all()
        return Company.objects.filter(id=self.request.user.userprofile.company_id)


# Lead Views
class LeadViewSet(viewsets.ModelViewSet):
    serializer_class = LeadSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        user = self.request.user
        if not user or not user.is_authenticated:
            return Lead.objects.none()

        user_profile = user.userprofile
        if user_profile.role == "SUPERADMIN":
            queryset = Lead.objects.all()
        else:
            queryset = Lead.objects.all()
            if user_profile.role not in ["ADMIN", "AM"]:
                queryset = queryset.filter(assigned_to=user.username)

        filter_type = self.request.query_params.get("filter")
        if filter_type == "walkins":
            walkin_lead_ids = (
                WalkIn.objects.all().values_list("lead_id", flat=True).distinct()
            )
            queryset = queryset.filter(id__in=walkin_lead_ids)

        return queryset.order_by("-created_at")

    def perform_create(self, serializer):
        serializer.save(company_id=self.request.user.userprofile.company_id)

    def create(self, request, *args, **kwargs):
        # Role-based validation
        user = request.user
        assigned_to = request.data.get("assigned_to")

        # Check if user is authenticated
        if user and user.is_authenticated:
            user_role = user.userprofile.role

            # If user is not ADMIN, SUPERADMIN or AM, they can only assign to themselves
            if user_role not in ["ADMIN", "SUPERADMIN", "AM"]:
                if assigned_to != user.username:
                    return Response(
                        {"error": "You are not authorized to assign leads to others."},
                        status=status.HTTP_403_FORBIDDEN,
                    )

        return super().create(request, *args, **kwargs)


# Student Views
class StudentViewSet(viewsets.ModelViewSet):
    def get_queryset(self):
        user = self.request.user
        if not user or not user.is_authenticated:
            return Student.objects.none()

        user_profile = user.userprofile
        if user_profile.role not in ["ADMIN", "SUPERADMIN", "AM", "CRO", "CRE"]:
            from rest_framework.exceptions import PermissionDenied

            raise PermissionDenied("Your role cannot access academics.")

        if user_profile.role == "SUPERADMIN":
            return Student.objects.filter(is_deleted=False).order_by("-admission_date")
        return Student.objects.filter(is_deleted=False).order_by("-admission_date")

    def perform_create(self, serializer):
        user = self.request.user
        user_profile = getattr(user, "userprofile", None)

        # Determine company_id and role safely across models
        from lead_enrollment.models import StaffUser

        staff = StaffUser.objects.filter(user_id=user.id).first()

        if staff:
            company_id = user_profile.company_id if user_profile else 0
            user_role = staff.role
        elif user_profile:
            company_id = user_profile.company_id
            user_role = user_profile.role
        else:
            company_id = 0
            user_role = "STAFF"

        if user_role not in ["ADMIN", "SUPERADMIN", "AM", "CRO", "CRE"]:
            from rest_framework.exceptions import PermissionDenied

            raise PermissionDenied("Your role cannot create admissions.")

        serializer.save(company_id=company_id)

    serializer_class = StudentSerializer
    permission_classes = [AllowAny]
    filter_backends = [filters.SearchFilter]
    search_fields = ["full_name", "student_id", "phone"]

    @action(detail=False, methods=["get"])
    def pending(self, request):
        # Fetch students who have not fully paid and are not deleted
        students = (
            Student.objects.filter(is_deleted=False)
            .exclude(fee_status="Fully Paid")
            .order_by("full_name")
        )

        results = []
        for s in students:
            # Calculate metrics similar to frontend logic
            # Use total_paid if available in installments JSON, else fallback to 0
            # (In a real system, we'd have a dedicated payment table for this)
            total_paid = (
                s.installments.get("total_paid", 0)
                if isinstance(s.installments, dict)
                else 0
            )

            # Formulate the response object for the table
            results.append(
                {
                    "id": s.id,
                    "student_id": s.student_id,
                    "full_name": s.full_name,
                    "course": s.courses[0].get("course")
                    if s.courses and len(s.courses) > 0
                    else "N/A",
                    "total_fee": float(s.total_fee),
                    "discount": float(s.discount),
                    "final_fee": float(s.final_fee),
                    "paid_amount": float(total_paid),
                    "pending_amount": float(s.final_fee) - float(total_paid),
                    "fee_status": s.fee_status,
                }
            )

        return Response(results)

    def create(self, request, *args, **kwargs):
        from django.http import JsonResponse
        from lead_enrollment.models import StaffUser, UserProfile

        user = request.user
        if not user or not user.is_authenticated:
            return JsonResponse({"error": "Authentication required."}, status=401)

        tenant_db = getattr(request, "tenant_db", "default")
        staff = StaffUser.objects.filter(user_id=user.id).first()

        if staff:
            user_role = staff.role
            company_id = (
                user.userprofile.company_id if hasattr(user, "userprofile") else 0
            )
        else:
            try:
                user_role = user.userprofile.role
                company_id = user.userprofile.company_id
            except UserProfile.DoesNotExist:
                user_role = "STAFF"
                company_id = 0

        if user_role not in ["ADMIN", "SUPERADMIN", "AM", "CRO", "CRE"]:
            return JsonResponse(
                {"error": "Your role is not authorized to create students."}, status=403
            )

        # Backend Validation for Required Custom Fields
        from lead_enrollment.models import CustomAdmissionField

        active_required_fields = CustomAdmissionField.objects.filter(
            is_active=True, required=True
        )

        custom_fields_raw = request.data.get("custom_fields", "{}")
        if isinstance(custom_fields_raw, str):
            import json

            try:
                custom_fields_input = json.loads(custom_fields_raw)
            except:
                custom_fields_input = {}
        else:
            custom_fields_input = (
                custom_fields_raw if isinstance(custom_fields_raw, dict) else {}
            )

        for field in active_required_fields:
            val = custom_fields_input.get(field.field_name)
            # Check if file exists for File Upload types
            has_file = request.FILES.get(f"custom_file_{field.field_name}")

            if field.field_type == "File Upload":
                if not has_file and not val:
                    return JsonResponse(
                        {"error": f"Custom field '{field.field_label}' is required."},
                        status=400,
                    )
            else:
                if val is None or (isinstance(val, str) and not val.strip()):
                    return JsonResponse(
                        {"error": f"Custom field '{field.field_label}' is required."},
                        status=400,
                    )

        # Step 8: LOG ADMISSION CREATED
        response = super().create(request, *args, **kwargs)
        if response.status_code == 201:
            student_id_val = response.data.get("student_id", "N/A")
            student_pk = response.data.get("id")
            student = Student.objects.get(pk=student_pk)

            # Handle Custom Fields
            custom_fields_data = request.data.get("custom_fields")
            if custom_fields_data:
                import json

                if isinstance(custom_fields_data, str):
                    try:
                        custom_fields_data = json.loads(custom_fields_data)
                    except:
                        custom_fields_data = {}

                if isinstance(custom_fields_data, dict):
                    for field_name, value in custom_fields_data.items():
                        field_obj = (
                            CustomAdmissionField.objects
                            .filter(field_name=field_name, is_active=True)
                            .first()
                        )
                        if field_obj:
                            final_value = str(value)
                            if value == "__FILE__":
                                uploaded_file = request.FILES.get(
                                    f"custom_file_{field_name}"
                                )
                                if uploaded_file:
                                    from django.core.files.storage import (
                                        default_storage,
                                    )
                                    import os

                                    # Save to student_files/tenant/field_name_filename
                                    file_path = f"student_files/{tenant_db}/{student.student_id}_{field_name}_{uploaded_file.name}"
                                    path = default_storage.save(
                                        file_path, uploaded_file
                                    )
                                    final_value = path

                            StudentCustomFieldValue.objects.create(
                                student=student, field=field_obj, value=final_value
                            )

            log_activity(user, "ADMISSION CREATED", f"New admission {student_id_val}")

            # Step 7: AUTO-CREATE INVOICE on Admission
            try:
                student_pk = response.data.get("id")
                student = Student.objects.get(pk=student_pk)
                year = datetime.now().year
                count = Invoice.objects.filter(company_id=company_id).count() + 1
                invoice_no = f"INV{year}{count:04d}"
                Invoice.objects.create(
                    student=student,
                    invoice_number=invoice_no,
                    total_fee=int(student.total_fee),
                    discount=int(student.discount),
                    final_fee=int(student.final_fee),
                    company_id=company_id,
                )
            except Exception as e:
                print(f"[Invoice] Auto-create failed: {e}")

            return JsonResponse(
                {
                    "message": "Admission completed successfully",
                    "student_id": student_id_val,
                    "data": response.data,
                },
                status=201,
            )

        return JsonResponse(
            {"error": "Failed to create admission", "details": response.data},
            status=400,
        )

    def update(self, request, *args, **kwargs):
        # ROLE-BASED VALIDATION: Only ADMIN/SUPERADMIN can update students
        user = request.user
        if user and user.is_authenticated:
            user_role = user.userprofile.role
            if user_role not in ["ADMIN", "SUPERADMIN"]:
                return Response(
                    {"error": "Only ADMIN is authorized to update students."},
                    status=status.HTTP_403_FORBIDDEN,
                )
        else:
            return Response(
                {"error": "Authentication required."},
                status=status.HTTP_401_UNAUTHORIZED,
            )
        return super().update(request, *args, **kwargs)

    def perform_update(self, serializer):
        instance = self.get_object()

        # Fields to track for history
        fields_to_track = [
            "full_name",
            "email",
            "phone",
            "dob",
            "gender",
            "blood_group",
            "batch",
            "address",
            "education_level",
            "college_name",
            "specialization",
            "year_of_passing",
            "marks",
            "total_fee",
            "discount",
            "final_fee",
            "payment_scheme",
            "payment_mode",
            "assigned_to",
            "lead_source",
        ]

        old_values = {}
        for f in fields_to_track:
            old_values[f] = getattr(instance, f)

        # Also track special items
        old_values["paid_amount"] = (
            instance.installments.get("total_paid", 0)
            if isinstance(instance.installments, dict)
            else 0
        )
        old_values["course"] = (
            instance.courses[0].get("course")
            if instance.courses and len(instance.courses) > 0
            else "N/A"
        )

        # Execute update
        company_id = self.request.user.userprofile.company_id
        updated_instance = serializer.save(company_id=company_id)

        # Compare and log changes
        user = self.request.user
        username = user.username if user and user.is_authenticated else "ADMIN"

        field_display_names = {
            "full_name": "Name",
            "email": "Email",
            "phone": "Phone",
            "dob": "DOB",
            "gender": "Gender",
            "blood_group": "Blood Group",
            "batch": "Batch",
            "address": "Address",
            "education_level": "Edu Level",
            "college_name": "College",
            "specialization": "Specialization",
            "year_of_passing": "Year",
            "marks": "Marks",
            "total_fee": "Total Fee",
            "discount": "Discount",
            "final_fee": "Final Fee",
            "payment_scheme": "Payment Type",
            "payment_mode": "Payment Mode",
            "assigned_to": "Assigned To",
            "lead_source": "Lead Source",
            "paid_amount": "Paid Amount",
            "course": "Course",
        }

        for field, old_val in old_values.items():
            new_val = None
            if field == "paid_amount":
                new_val = (
                    updated_instance.installments.get("total_paid", 0)
                    if isinstance(updated_instance.installments, dict)
                    else 0
                )
            elif field == "course":
                new_val = (
                    updated_instance.courses[0].get("course")
                    if updated_instance.courses and len(updated_instance.courses) > 0
                    else "N/A"
                )
            elif field in ["total_fee", "discount", "final_fee"]:
                new_val = getattr(updated_instance, field)
                if float(old_val or 0) == float(new_val or 0):
                    continue
            elif field == "payment_scheme":
                new_val = updated_instance.payment_scheme
                old_val_log = "Term-wise" if old_val == "Installment" else "One-time"
                new_val_log = "Term-wise" if new_val == "Installment" else "One-time"
                if old_val_log != new_val_log:
                    StudentChangeLog.objects.create(
                        student=updated_instance,
                        field_name=field_display_names[field],
                        old_value=old_val_log,
                        new_value=new_val_log,
                        changed_by=username,
                        company_id=company_id,
                    )
                continue
            else:
                new_val = getattr(updated_instance, field)

            if str(old_val) != str(new_val) and str(old_val or "") != str(
                new_val or ""
            ):
                StudentChangeLog.objects.create(
                    student=updated_instance,
                    field_name=field_display_names.get(field, field),
                    old_value=str(old_val) if old_val is not None else "N/A",
                    new_value=str(new_val) if new_val is not None else "N/A",
                    changed_by=username,
                    company_id=company_id,
                )

        # Step 9: LOG PAYMENT UPDATE
        if "paid_amount" in old_values:
            new_paid = (
                updated_instance.installments.get("total_paid", 0)
                if isinstance(updated_instance.installments, dict)
                else 0
            )
            if float(old_values["paid_amount"]) != float(new_paid):
                log_activity(
                    user,
                    "PAYMENT UPDATED",
                    f"Payment updated for {updated_instance.student_id}",
                )

                # Step 7: AUTO-CREATE RECEIPT on Payment
                try:
                    paid_amount_diff = float(new_paid) - float(
                        old_values["paid_amount"]
                    )
                    if paid_amount_diff > 0:
                        year = datetime.now().year
                        count = (
                            Receipt.objects.filter(company_id=company_id).count() + 1
                        )
                        receipt_no = f"REC{year}{count:04d}"
                        payment_mode = (
                            getattr(updated_instance, "payment_mode", "CASH") or "CASH"
                        )
                        Receipt.objects.create(
                            student=updated_instance,
                            receipt_number=receipt_no,
                            amount_paid=int(paid_amount_diff),
                            payment_mode=payment_mode,
                            company_id=company_id,
                        )
                except Exception as e:
                    print(f"[Receipt] Auto-create failed: {e}")

        # Step 6: LOG STUDENT EDIT
        log_activity(
            user, "STUDENT UPDATED", f"Edited student {updated_instance.student_id}"
        )

    def partial_update(self, request, *args, **kwargs):
        # ROLE-BASED VALIDATION: Only ADMIN/SUPERADMIN can update students
        user = request.user
        if user and user.is_authenticated:
            user_role = user.userprofile.role
            if user_role not in ["ADMIN", "SUPERADMIN"]:
                return Response(
                    {"error": "Only ADMIN is authorized to update students."},
                    status=status.HTTP_403_FORBIDDEN,
                )
        else:
            return Response(
                {"error": "Authentication required."},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        # CERTIFICATE VALIDATION: Block certificate application if fees not fully paid
        if request.data.get("certificate_status") == "Applied":
            student = self.get_object()
            # Check fee_status field
            if student.fee_status != "Fully Paid":
                return Response(
                    {
                        "error": "Certificate cannot be applied until full payment is completed."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
            # Double-check: compute pending amount from installments
            total_paid = 0
            if isinstance(student.installments, dict):
                total_paid = float(student.installments.get("total_paid", 0))
            pending_balance = float(student.final_fee) - total_paid
            if pending_balance > 0.01:  # Allow small floating-point tolerance
                return Response(
                    {
                        "error": "Certificate cannot be applied until full payment is completed."
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )

        return super().partial_update(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        # ROLE-BASED VALIDATION: Only ADMIN/SUPERADMIN can delete
        user = request.user
        if user and user.is_authenticated:
            user_role = user.userprofile.role
            if user_role not in ["ADMIN", "SUPERADMIN"]:
                return Response(
                    {"error": "Only ADMIN is authorized to delete students."},
                    status=status.HTTP_403_FORBIDDEN,
                )
        else:
            # Fallback if somehow not authenticated but reached here
            return Response(
                {"error": "Authentication required."},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        instance = self.get_object()
        student_id = instance.student_id
        instance.is_deleted = True
        instance.save()

        # Step 7: LOG STUDENT DELETE
        log_activity(user, "STUDENT DELETED", f"Deleted student {student_id}")

        return Response(
            {"message": "Student deleted successfully."}, status=status.HTTP_200_OK
        )

    @action(detail=True, methods=["get"])
    def history(self, request, pk=None):
        # ROLE-BASED VALIDATION
        user = request.user
        if (
            not user
            or not user.is_authenticated
            or user.userprofile.role not in ["ADMIN", "SUPERADMIN"]
        ):
            return Response(
                {"error": "Only ADMIN is authorized to view history."},
                status=status.HTTP_403_FORBIDDEN,
            )

        student = self.get_object()
        limit = int(request.query_params.get("limit", 50))
        offset = int(request.query_params.get("offset", 0))

        logs = StudentChangeLog.objects.filter(student=student).order_by("-changed_at")[
            offset : offset + limit
        ]

        data = [
            {
                "id": l.id,
                "field": l.field_name,
                "old": l.old_value,
                "new": l.new_value,
                "user": l.changed_by,
                "date": l.changed_at.strftime("%Y-%m-%d %H:%M"),
            }
            for l in logs
        ]

        return Response(data)

    @action(detail=True, methods=["post"])
    def undo(self, request, pk=None):
        # ROLE-BASED VALIDATION
        user = request.user
        if (
            not user
            or not user.is_authenticated
            or user.userprofile.role not in ["ADMIN", "SUPERADMIN"]
        ):
            return Response(
                {"error": "Only ADMIN is authorized to undo changes."},
                status=status.HTTP_403_FORBIDDEN,
            )

        student = self.get_object()
        history_id = request.data.get("history_id")

        try:
            log = StudentChangeLog.objects.get(id=history_id, student=student)
        except StudentChangeLog.DoesNotExist:
            return Response({"error": "History record not found."}, status=404)

        field_name = log.field_name
        old_value = log.old_value

        display_to_field = {
            "Name": "full_name",
            "Email": "email",
            "Phone": "phone",
            "DOB": "dob",
            "Gender": "gender",
            "Blood Group": "blood_group",
            "Batch": "batch",
            "Address": "address",
            "Edu Level": "education_level",
            "College": "college_name",
            "Specialization": "specialization",
            "Year": "year_of_passing",
            "Marks": "marks",
            "Total Fee": "total_fee",
            "Discount": "discount",
            "Final Fee": "final_fee",
            "Payment Type": "payment_scheme",
            "Payment Mode": "payment_mode",
            "Assigned To": "assigned_to",
            "Lead Source": "lead_source",
            "Paid Amount": "paid_amount",
            "Course": "course",
        }

        field = display_to_field.get(field_name)
        if not field:
            return Response(
                {"error": f"Undo not supported for {field_name}"}, status=400
            )

        # Revert student status
        if field == "paid_amount":
            if not student.installments:
                student.installments = {}
            student.installments["total_paid"] = float(old_value)
        elif field == "course":
            if student.courses and len(student.courses) > 0:
                student.courses[0]["course"] = old_value
            else:
                student.courses = [
                    {"id": 1, "category": "General", "course": old_value, "fee": 0}
                ]
        elif field == "payment_scheme":
            student.payment_scheme = (
                "Installment" if old_value == "Term-wise" else "One-time"
            )
        else:
            setattr(student, field, old_value if old_value != "N/A" else None)

        student.save()

        # Log the undo as a new record
        StudentChangeLog.objects.create(
            student=student,
            field_name=field_name,
            old_value=log.new_value,
            new_value=log.old_value,
            changed_by=user.username,
        )

        return Response({"message": "Change reverted successfully"})

    @action(detail=True, methods=["post"], url_path="upload-photo")
    def upload_photo(self, request, pk=None):
        student = self.get_object()
        photo = request.FILES.get("photo")
        if not photo:
            return Response({"error": "No photo provided."}, status=400)

        ext = photo.name.split(".")[-1].lower()
        if ext not in ["jpg", "jpeg"]:
            return Response({"error": "Only jpg, jpeg allowed."}, status=400)

        if photo.size > 25600:
            return Response({"error": "File must be below 25 KB"}, status=400)

        # Delete old photo if it exists
        if student.student_photo and os.path.isfile(student.student_photo.path):
            try:
                os.remove(student.student_photo.path)
            except Exception as e:
                print(f"Error removing old photo: {e}")

        student.student_photo = photo
        student.save()
        return Response(StudentSerializer(student, context={"request": request}).data)

    @action(detail=True, methods=["post"], url_path="upload-id")
    def upload_id(self, request, pk=None):
        student = self.get_object()
        id_proof = request.FILES.get("id_proof")
        id_proof_type = request.data.get("id_proof_type")

        if not id_proof:
            return Response({"error": "No ID proof provided."}, status=400)

        ext = id_proof.name.split(".")[-1].lower()
        if ext not in ["pdf", "jpg", "jpeg"]:
            return Response({"error": "Only pdf, jpg, jpeg allowed."}, status=400)

        if id_proof.size > 25600:
            return Response({"error": "File must be below 25 KB"}, status=400)

        # Delete old ID proof if it exists
        if student.id_proof and os.path.isfile(student.id_proof.path):
            try:
                os.remove(student.id_proof.path)
            except Exception as e:
                print(f"Error removing old ID proof: {e}")

        student.id_proof = id_proof
        if id_proof_type:
            student.id_proof_type = id_proof_type
        student.save()
        return Response(StudentSerializer(student, context={"request": request}).data)


# LeadRemark Views
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_student_by_roll_no(request, roll_no):
    from .models import Student, StudentCustomFieldValue
    from .serializers import StudentSerializer

    try:
        student = Student.objects.get(student_id=roll_no, is_deleted=False)
        # Calculate paid amount (logic consistent with other parts)
        total_paid = (
            student.installments.get("total_paid", 0)
            if isinstance(student.installments, dict)
            else 0
        )

        serializer = StudentSerializer(student)
        data = serializer.data
        data["paid_amount"] = float(total_paid)
        data["pending_amount"] = float(student.final_fee) - float(total_paid)

        # Fetch custom field values for this student
        custom_values = StudentCustomFieldValue.objects.filter(
            student=student
        ).select_related("field")
        data["custom_fields"] = [
            {
                "field_label": cv.field.field_label,
                "field_name": cv.field.field_name,
                "value": cv.value,
                "tab": cv.field.tab_name,
                "field_type": cv.field.field_type,
            }
            for cv in custom_values
            if cv.field.is_active
        ]

        return Response(data)
    except Student.DoesNotExist:
        return Response(
            {"error": "Student not found"}, status=status.HTTP_404_NOT_FOUND
        )


@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def update_custom_field_value(request, id):
    from .models import StudentCustomFieldValue

    try:
        tenant_db = getattr(request, "tenant_db", "default")
        obj = StudentCustomFieldValue.objects.get(id=id)

        # Security check: can only edit own company's students if not Superadmin
        # (Assuming Student model has company_id, and obj.student points to Student)
        user = request.user
        if not user.userprofile.role == "SUPERADMIN":
            if obj.student.company_id != user.userprofile.company_id:
                return Response(
                    {"error": "Unauthorized access to this student's data."}, status=403
                )

        new_value = request.data.get("value")
        if new_value is not None:
            obj.value = str(new_value)
            obj.save(using=tenant_db)

            # Log the change in history if possible
            # (Skipping for now to keep it simple, but good practice)

            return Response({"status": "success", "value": obj.value})
        return Response({"error": "No value provided"}, status=400)
    except StudentCustomFieldValue.DoesNotExist:
        return Response({"error": "Value record not found"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_student_by_roll(request, roll_no):
    from .models import Student

    user = request.user
    # Role validation: only ADMIN/SUPERADMIN can delete
    if not hasattr(user, "userprofile") or user.userprofile.role not in [
        "ADMIN",
        "SUPERADMIN",
    ]:
        return Response(
            {"error": "Only ADMIN is authorized to delete students."}, status=403
        )

    try:
        student = Student.objects.get(student_id=roll_no, is_deleted=False)
        student.is_deleted = True
        student.save()

        log_activity(user, "STUDENT DELETED", f"Deleted student {roll_no}")

        return Response(
            {"status": "success", "message": "Student deleted successfully"}, status=200
        )
    except Student.DoesNotExist:
        return Response({"error": "Student not found"}, status=404)


@api_view(["PUT"])
@permission_classes([IsAuthenticated])
def update_student_by_roll(request, roll_no):
    from .models import Student

    user = request.user
    # Role validation
    if not hasattr(user, "userprofile") or user.userprofile.role != "ADMIN":
        return Response(
            {"error": "Only ADMIN is authorized to update student details."}, status=403
        )

    try:
        student = Student.objects.get(student_id=roll_no, is_deleted=False)
        data = request.data

        # Update primary fields
        if "name" in data:
            student.full_name = data["name"]
        if "email" in data:
            student.email = data["email"]
        if "phone" in data:
            student.phone = data["phone"]
        if "address" in data:
            student.address = data["address"]
        if "education" in data:
            student.education_level = data["education"]
        if "college" in data:
            student.college_name = data["college"]
        if "specialization" in data:
            student.specialization = data["specialization"]
        if "year_passing" in data:
            student.year_of_passing = data["year_passing"]
        if "guardian_name" in data:
            student.guardian1_name = data["guardian_name"]
        if "guardian_phone" in data:
            student.guardian1_phone = data["guardian_phone"]
        if "batch" in data:
            student.batch = data["batch"]
        if "assigned_staff" in data:
            student.assigned_to = data["assigned_staff"]

        # Files
        if "student_photo" in request.FILES:
            student.student_photo = request.FILES["student_photo"]
        if "id_proof" in request.FILES:
            student.id_proof = request.FILES["id_proof"]
        if "id_proof_type" in data:
            student.id_proof_type = data["id_proof_type"]

        # Fee & Payment Logic
        try:
            total_fee = float(data.get("total_fee", student.total_fee) or 0)
            discount = float(data.get("discount", student.discount) or 0)
        except (ValueError, TypeError):
            total_fee = float(student.total_fee or 0)
            discount = float(student.discount or 0)

        # Calculate Effective Fee (final_fee)
        student.total_fee = total_fee
        student.discount = discount
        student.final_fee = total_fee - discount

        # Update Paid Amount in installments
        if "paid_amount" in data:
            try:
                paid_val = float(data["paid_amount"] or 0)
                if not student.installments:
                    student.installments = {}
                student.installments["total_paid"] = paid_val
            except (ValueError, TypeError):
                pass

        # Recalculate fee status
        total_paid = (
            float(student.installments.get("total_paid", 0))
            if isinstance(student.installments, dict)
            else 0
        )

        if total_paid >= float(student.final_fee):
            student.fee_status = "Fully Paid"
        elif total_paid > 0:
            student.fee_status = "Partially Paid"
        else:
            student.fee_status = "Pending"

        # Course update logic
        if "course" in data:
            new_course_name = data["course"]
            if student.courses and len(student.courses) > 0:
                if isinstance(student.courses[0], dict):
                    student.courses[0]["course"] = new_course_name
            else:
                student.courses = [
                    {
                        "id": 1,
                        "category": "General",
                        "course": new_course_name,
                        "fee": 0,
                    }
                ]

        student.save()

        # Custom Fields Logic
        custom_fields_data = data.get("custom_fields")
        if custom_fields_data:
            import json

            if isinstance(custom_fields_data, str):
                try:
                    custom_fields_data = json.loads(custom_fields_data)
                except:
                    custom_fields_data = {}

            if isinstance(custom_fields_data, dict):
                tenant_db = getattr(request, "tenant_db", "default")
                from .models import CustomAdmissionField, StudentCustomFieldValue

                for field_name, value in custom_fields_data.items():
                    field_obj = (
                        CustomAdmissionField.objects
                        .filter(field_name=field_name, is_active=True)
                        .first()
                    )
                    if field_obj:
                        final_value = str(value)
                        if value == "__FILE__":
                            uploaded_file = request.FILES.get(
                                f"custom_file_{field_name}"
                            )
                            if uploaded_file:
                                from django.core.files.storage import default_storage

                                # Save to student_files/tenant/field_name_filename
                                file_path = f"student_files/{tenant_db}/{student.student_id}_{field_name}_{uploaded_file.name}"
                                path = default_storage.save(file_path, uploaded_file)
                                final_value = path
                            else:
                                # If no new file uploaded, keep existing value
                                existing = (
                                    StudentCustomFieldValue.objects
                                    .filter(student=student, field=field_obj)
                                    .first()
                                )
                                if existing:
                                    final_value = existing.value

                        StudentCustomFieldValue.objects.using(
                            tenant_db
                        ).update_or_create(
                            student=student,
                            field=field_obj,
                            defaults={"value": final_value},
                        )

        return Response(
            {"status": "success", "message": "Student details updated successfully"}
        )
    except Student.DoesNotExist:
        return Response({"error": "Student not found"}, status=404)
    except Exception as e:
        return Response({"error": str(e)}, status=400)


@csrf_exempt
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def confirm_admission(request):
    """
    Standalone endpoint for confirming admission.
    Handles both JSON and FormData requests.
    """
    import json

    try:
        # Standard DRF request.data handles both JSON and FormData parsing automatically
        # when using @api_view(['POST']).
        data = request.data

        # Use StudentViewSet's logic to handle the heavy lifting
        # while wrapping it in the requested try/except JSON response pattern
        viewset = StudentViewSet()
        viewset.request = request
        viewset.format_kwarg = None
        viewset.action = "create"

        # Call the existing StudentViewSet.create logic which we've already
        # optimized for performance and role safety.
        response = viewset.create(request)

        # Ensure we return the response directly as it's already a JsonResponse
        # from our previous optimization.
        return response

    except Exception as e:
        print(f"[Confirm Admission] Crash: {str(e)}")
        return JsonResponse(
            {
                "status": "error",
                "error": str(e),
                "message": "Internal server error during admission confirmation",
            },
            status=500,
        )


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def save_admission(request):
    """
    Requested endpoint to save admission data including custom fields.
    """
    viewset = StudentViewSet()
    viewset.request = request
    viewset.format_kwarg = None
    viewset.action = "create"
    return viewset.create(request)


class CustomAdmissionFieldViewSet(viewsets.ModelViewSet):
    serializer_class = CustomAdmissionFieldSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        return CustomAdmissionField.objects.all().order_by("created_at")

    def perform_create(self, serializer):
        serializer.save()


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_custom_fields(request):
    try:
        from .models import CustomAdmissionField

        if request.GET.get("all") == "true":
            fields = CustomAdmissionField.objects.all().order_by("order")
        else:
            fields = CustomAdmissionField.objects.filter(is_active=True).order_by(
                "order"
            )

        data = [
            {
                "id": f.id,
                "tab_name": f.tab_name,
                "field_label": f.field_label,
                "field_name": f.field_name,
                "field_type": f.field_type,
                "placeholder": f.placeholder,
                "required": f.required,
                "is_active": f.is_active,
                "order": f.order,
                "options": f.options or [],
            }
            for f in fields
        ]

        return Response(data)

    except Exception as e:
        import traceback

        error_message = str(e)
        error_trace = traceback.format_exc()
        print(f"Fetch error: {error_message}")
        print(error_trace)

        return Response({"error": error_message, "trace": error_trace}, status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def add_custom_field(request):
    """
    Stabilized endpoint to add custom admission fields with error handling.
    """
    if request.user.userprofile.role not in ["ADMIN", "SUPERADMIN"]:
        return Response({"message": "Unauthorized"}, status=status.HTTP_403_FORBIDDEN)

    try:
        data = request.data

        field = CustomAdmissionField.objects.create(
            tab_name=data.get("tab_name"),
            field_label=data.get("field_label"),
            field_name=data.get("field_name")
            or data.get("field_label", "").lower().replace(" ", "_"),
            field_type=data.get("field_type"),
            placeholder=data.get("placeholder", ""),
            required=data.get("required", False),
            options=data.get("options", []),
            is_active=True,
        )

        return Response(
            {"message": "Custom field created successfully", "field_id": field.id}
        )

    except Exception as e:
        import traceback

        print("Add custom field error:", str(e))
        print(traceback.format_exc())
        return Response({"message": str(e)}, status=500)


@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def toggle_field_status(request, field_id):
    """
    Toggle the active status of a custom admission field.
    """
    if request.user.userprofile.role not in ["ADMIN", "SUPERADMIN"]:
        return Response({"message": "Unauthorized"}, status=status.HTTP_403_FORBIDDEN)

    try:
        from .models import CustomAdmissionField

        field = CustomAdmissionField.objects.get(id=field_id)
        field.is_active = not field.is_active
        field.save()

        print(f"✅ Toggle field {field_id}: is_active = {field.is_active}")  # DEBUG

        return Response({"status": "updated", "is_active": field.is_active})
    except Exception as e:
        print(f"❌ Toggle error: {str(e)}")
        return Response({"error": str(e)}, status=500)


@api_view(["PATCH"])
@permission_classes([IsAuthenticated])
def toggle_required_status(request, field_id):
    """
    Toggle the required status of a custom admission field.
    """
    if request.user.userprofile.role not in ["ADMIN", "SUPERADMIN"]:
        return Response({"message": "Unauthorized"}, status=status.HTTP_403_FORBIDDEN)

    try:
        from .models import CustomAdmissionField

        field = CustomAdmissionField.objects.get(id=field_id)
        field.required = not field.required
        field.save()

        return Response({"status": "updated", "required": field.required})
    except Exception as e:
        return Response({"error": str(e)}, status=500)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def update_field_order(request):
    """
    Update the order of custom admission fields (Drag & Drop).
    """
    if request.user.userprofile.role not in ["ADMIN", "SUPERADMIN"]:
        return Response({"message": "Unauthorized"}, status=status.HTTP_403_FORBIDDEN)

    try:
        from .models import CustomAdmissionField

        items = request.data  # Expecting list of {id, order}

        for item in items:
            CustomAdmissionField.objects.filter(id=item.get("id")).update(
                order=item.get("order")
            )

        return Response({"message": "Field order updated successfully"})
    except Exception as e:
        return Response({"message": str(e)}, status=500)


class LeadRemarkViewSet(viewsets.ModelViewSet):
    queryset = LeadRemark.objects.all()
    serializer_class = LeadRemarkSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        user = self.request.user
        if not user or not user.is_authenticated:
            return LeadRemark.objects.none()

        user_profile = user.userprofile
        if user_profile.role == "SUPERADMIN":
            queryset = LeadRemark.objects.all()
        else:
            queryset = LeadRemark.objects.all()

        lead_id = self.request.query_params.get("lead_id")
        if lead_id:
            queryset = queryset.filter(lead_id=lead_id)
        return queryset

    def perform_create(self, serializer):
        serializer.save(company_id=self.request.user.userprofile.company_id)


# WalkIn Views
class WalkInViewSet(viewsets.ModelViewSet):
    queryset = WalkIn.objects.all()
    serializer_class = WalkInSerializer
    permission_classes = [AllowAny]

    def get_queryset(self):
        user = self.request.user
        if not user or not user.is_authenticated:
            return WalkIn.objects.none()

        user_profile = user.userprofile
        if user_profile.role == "SUPERADMIN":
            queryset = WalkIn.objects.all()
        else:
            queryset = WalkIn.objects.all()

        lead_id = self.request.query_params.get("lead_id")
        recorded_by = self.request.query_params.get("recorded_by")
        if lead_id:
            queryset = queryset.filter(lead_id=lead_id)
        if recorded_by:
            queryset = queryset.filter(recorded_by=recorded_by)
        return queryset

    def perform_create(self, serializer):
        serializer.save(company_id=self.request.user.userprofile.company_id)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def dashboard_stats(request):
    db = getattr(request, "tenant_db", "default")

    # 1. Error handling if tenant_db is missing or default (for non-superadmins)
    is_superadmin = request.user.userprofile.role == "SUPERADMIN"
    if not is_superadmin and db == "default":
        return Response({"error": "Tenant database not assigned"}, status=400)

    # 2. Base Queries using the tenant database
    leads_qs = Lead.objects.all()
    walkins_qs = WalkIn.objects.all()
    students_qs = Student.objects.filter(is_deleted=False)

    # Basic KPI counts
    total_leads = leads_qs.count()
    total_walk_ins = walkins_qs.values("lead").distinct().count()
    total_registered = students_qs.count()

    # Financial data from Students
    total_collection = students_qs.aggregate(Sum("final_fee"))["final_fee__sum"] or 0
    total_pending = (
        students_qs.filter(fee_status="Pending").aggregate(Sum("final_fee"))[
            "final_fee__sum"
        ]
        or 0
    )
    total_paid = (
        students_qs.filter(fee_status="Fully Paid").aggregate(Sum("final_fee"))[
            "final_fee__sum"
        ]
        or 0
    )

    # Lead Sources Breakdown
    sources_data = (
        leads_qs.values("source").annotate(count=Count("source")).order_by("-count")
    )

    # Lead Acquisition Trend (last 7 days)
    today = timezone.now().date()
    seven_days_ago = today - timedelta(days=6)

    trend_data = []
    for i in range(7):
        date = seven_days_ago + timedelta(days=i)
        count = leads_qs.filter(created_at__date=date).count()
        trend_data.append(
            {"date": date.strftime("%a"), "count": count}  # Mon, Tue, etc.
        )

    return Response(
        {
            "kpis": {
                "total_leads": total_leads,
                "total_walk_ins": total_walk_ins,
                "total_registered": total_registered,
                "total_collection_formatted": f"₹ {total_collection / 100000:.1f}L"
                if total_collection >= 100000
                else f"₹ {total_collection}",
                "total_collection_raw": total_collection,
                "paid_amount": total_paid,
                "pending_amount": total_pending,
            },
            "sources": list(sources_data),
            "trend": trend_data,
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def tracker_stats(request):
    if request.user.userprofile.role not in ["ADMIN", "SUPERADMIN", "AM"]:
        return Response({"detail": "Permission denied."}, status=403)

    # Filters
    year = request.query_params.get("year", str(datetime.now().year))
    month = request.query_params.get("month")
    week = request.query_params.get("week")

    # Base QuerySets
    db = getattr(request, "tenant_db", "default")
    is_superadmin = request.user.userprofile.role == "SUPERADMIN"

    if not is_superadmin and db == "default":
        return Response({"error": "Tenant database not assigned"}, status=400)

    leads_qs = Lead.objects.all()
    students_qs = Student.objects.filter(is_deleted=False)
    walkins_qs = WalkIn.objects.all()

    try:
        year_num = int(year)
        leads_qs = leads_qs.filter(created_at__year=year_num)
        students_qs = students_qs.filter(admission_date__year=year_num)
        walkins_qs = walkins_qs.filter(created_at__year=year_num)

        if month:
            month_num = datetime.strptime(month, "%B").month

            if week and week.startswith("Week"):
                # Week selection: calculate date range within the month
                try:
                    week_num = int(week.split(" ")[1])
                    if week_num == 1:
                        start_day, end_day = 1, 7
                    elif week_num == 2:
                        start_day, end_day = 8, 14
                    elif week_num == 3:
                        start_day, end_day = 15, 21
                    elif week_num == 4:
                        start_day = 22
                        _, end_day = calendar.monthrange(year_num, month_num)

                    start_date = timezone.make_aware(
                        datetime(year_num, month_num, start_day, 0, 0, 0)
                    )
                    end_date = timezone.make_aware(
                        datetime(year_num, month_num, end_day, 23, 59, 59)
                    )

                    leads_qs = leads_qs.filter(created_at__range=(start_date, end_date))
                    # Students admission_date is a DateField, so we use .date()
                    students_qs = students_qs.filter(
                        admission_date__range=(start_date.date(), end_date.date())
                    )
                    walkins_qs = walkins_qs.filter(
                        created_at__range=(start_date, end_date)
                    )
                except (ValueError, IndexError):
                    # Fallback to month if week parsing fails
                    leads_qs = leads_qs.filter(created_at__month=month_num)
                    students_qs = students_qs.filter(admission_date__month=month_num)
                    walkins_qs = walkins_qs.filter(created_at__month=month_num)
            else:
                # Only month
                leads_qs = leads_qs.filter(created_at__month=month_num)
                students_qs = students_qs.filter(admission_date__month=month_num)
                walkins_qs = walkins_qs.filter(created_at__month=month_num)
    except ValueError:
        pass  # Invalid year or other numeric parsing error, fallback to un-filtered collections

    # Calculate Tracker KPIs
    total_leads = leads_qs.count()
    relevant_leads = leads_qs.exclude(status="Lost").exclude(status="Dead").count()
    walk_ins = walkins_qs.values("lead").distinct().count()

    # Financials for Tracker (based on Students created in this period)
    # Note: This might not track *payments* made in this period for old students,
    # but tracks revenue generated by *new* admissions in this period.
    amount_collected = students_qs.aggregate(Sum("final_fee"))["final_fee__sum"] or 0
    # Ideally should be based on payment transactions, but using Admission Final Fee as proxy for now
    # or strictly paid amount:
    # amount_collected = students_qs.filter(fee_status='Fully Paid').aggregate(Sum('final_fee'))['final_fee__sum'] or 0

    # Staff Performance
    # Group by assigned_to
    staff_data = []

    # Get all unique staff names from leads and students
    staff_names = set(leads_qs.values_list("assigned_to", flat=True).distinct()) | set(
        students_qs.values_list("assigned_to", flat=True).distinct()
    )

    for staff in staff_names:
        if not staff:
            continue

        staff_leads = leads_qs.filter(assigned_to=staff).count()
        staff_walkins = (
            walkins_qs.filter(recorded_by=staff).values("lead").distinct().count()
        )
        staff_admissions = students_qs.filter(assigned_to=staff).count()
        staff_revenue = (
            students_qs.filter(assigned_to=staff).aggregate(Sum("final_fee"))[
                "final_fee__sum"
            ]
            or 0
        )

        # Determine role (simple heuristic or need a User model join)
        role = "Staff"
        if "ADMIN" in staff.upper():
            role = "Admin"
        elif "CRE" in staff.upper():
            role = "CRE"
        elif "CRO" in staff.upper():
            role = "CRO"

        staff_data.append(
            {
                "name": staff,
                "role": role,
                "leads_assigned": staff_leads,
                "walk_ins": staff_walkins,
                "admissions": staff_admissions,
                "revenue": staff_revenue,
            }
        )

    return Response(
        {
            "kpis": {
                "total_leads": total_leads,
                "relevant_leads": relevant_leads,
                "walk_ins": walk_ins,
                "amount_collected": amount_collected,
                "amount_collected_formatted": f"₹ {amount_collected / 100000:.1f}L"
                if amount_collected >= 100000
                else f"₹ {amount_collected}",
            },
            "staff_performance": staff_data,
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_stats(request):
    if request.user.userprofile.role not in ["ADMIN", "SUPERADMIN", "AM"]:
        return Response({"detail": "Permission denied."}, status=403)

    filter_type = request.query_params.get("filter", "Month")  # Day, Week, Month, Year

    today = timezone.now().date()
    start_date = None

    if filter_type == "Day":
        start_date = today - timedelta(days=7)
    elif filter_type == "Week":
        start_date = today - timedelta(days=30)
    elif filter_type == "Month":
        start_date = today - timedelta(days=180)  # Last 6 months
    elif filter_type == "Year":
        start_date = today - timedelta(days=365 * 5)  # Last 5 years

    # Filter students based on period
    db = getattr(request, "tenant_db", "default")
    is_superadmin = request.user.userprofile.role == "SUPERADMIN"

    if not is_superadmin and db == "default":
        return Response({"error": "Tenant database not assigned"}, status=400)

    students_qs = Student.objects.filter(is_deleted=False)

    if start_date:
        students_qs = students_qs.filter(admission_date__gte=start_date)

    # Basic Financial KPIs based on filtered period
    total_collected = (
        students_qs.filter(fee_status="Fully Paid").aggregate(Sum("final_fee"))[
            "final_fee__sum"
        ]
        or 0
    )
    total_revenue = students_qs.aggregate(Sum("final_fee"))["final_fee__sum"] or 0

    # Calculate Pending Details
    pending_students_qs = students_qs.exclude(fee_status="Fully Paid")
    pending_list = []
    total_pending = 0

    for s in pending_students_qs:
        total_paid = (
            s.installments.get("total_paid", 0)
            if isinstance(s.installments, dict)
            else 0
        )
        balance = float(s.final_fee) - float(total_paid)

        if balance > 0:
            total_pending += balance
            pending_list.append(
                {
                    "id": s.id,
                    "student_id": s.student_id,
                    "full_name": s.full_name,
                    "course": s.courses[0].get("course")
                    if s.courses and len(s.courses) > 0
                    else "N/A",
                    "total_fee": float(s.total_fee),
                    "discount": float(s.discount),
                    "final_fee": float(s.final_fee),
                    "paid_amount": float(total_paid),
                    "pending_amount": balance,
                    "fee_status": s.fee_status,
                }
            )

    pending_count = len(pending_list)

    collected_list = []
    revenue_list = []
    course_revenue = {}
    staff_revenue = {}

    for s in students_qs:
        fee = float(s.final_fee)
        c_name = (
            s.courses[0].get("course")
            if isinstance(s.courses, list) and len(s.courses) > 0
            else "N/A"
        )

        revenue_list.append(
            {
                "id": s.id,
                "student_id": s.student_id,
                "full_name": s.full_name,
                "course": c_name,
                "amount": fee,
                "status": s.fee_status,
                "date": s.admission_date,
            }
        )

        if s.fee_status == "Fully Paid":
            collected_list.append(
                {
                    "id": s.id,
                    "student_id": s.student_id,
                    "full_name": s.full_name,
                    "course": c_name,
                    "amount": fee,
                    "date": s.admission_date,
                }
            )

        course_revenue[c_name] = course_revenue.get(c_name, 0) + fee
        staff = s.assigned_to or "Unassigned"
        staff_revenue[staff] = staff_revenue.get(staff, 0) + fee

    course_breakdown = [{"course": k, "amount": v} for k, v in course_revenue.items()]
    staff_breakdown = [{"staff": k, "amount": v} for k, v in staff_revenue.items()]
    course_breakdown.sort(key=lambda x: x["amount"], reverse=True)
    staff_breakdown.sort(key=lambda x: x["amount"], reverse=True)
    # Chart Data Construction (Keep existing logic but ensure it uses today/timezone)
    labels = []
    data_points = []

    if filter_type == "Day":
        # Last 7 days
        for i in range(6, -1, -1):
            date = today - timedelta(days=i)
            labels.append(date.strftime("%d %b"))
            daily_rev = (
                Student.objects
                .filter(admission_date=date, is_deleted=False)
                .aggregate(Sum("final_fee"))["final_fee__sum"]
                or 0
            )
            data_points.append(daily_rev)

    elif filter_type == "Week":
        for i in range(3, -1, -1):
            date_end = today - timedelta(days=i * 7)
            date_start = date_end - timedelta(days=7)
            labels.append(f"W{(today - date_end).days // 7 + 1}")
            week_rev = (
                Student.objects
                .filter(
                    admission_date__gt=date_start,
                    admission_date__lte=date_end,
                    is_deleted=False,
                )
                .aggregate(Sum("final_fee"))["final_fee__sum"]
                or 0
            )
            data_points.append(week_rev)

    elif filter_type == "Month":
        # Last 6 months
        for i in range(5, -1, -1):
            date = today - timedelta(days=i * 30)
            labels.append(date.strftime("%b"))
            month_rev = (
                Student.objects
                .filter(
                    admission_date__month=date.month,
                    admission_date__year=date.year,
                    is_deleted=False,
                )
                .aggregate(Sum("final_fee"))["final_fee__sum"]
                or 0
            )
            data_points.append(month_rev)

    elif filter_type == "Year":
        # Last 5 years
        for i in range(4, -1, -1):
            year = today.year - i
            labels.append(str(year))
            year_rev = (
                Student.objects
                .filter(admission_date__year=year, is_deleted=False)
                .aggregate(Sum("final_fee"))["final_fee__sum"]
                or 0
            )
            data_points.append(year_rev)

    # Recent Transactions (Simulated using recent admissions)
    recent_students = (
        Student.objects.filter(is_deleted=False).order_by("-created_at")[:5]
    )
    transactions = []
    for s in recent_students:
        transactions.append(
            {
                "student": s.full_name,
                "amount": f"₹ {s.final_fee:,}",
                "date": s.admission_date,
            }
        )

    return Response(
        {
            "kpis": {
                "collected": float(total_collected),
                "collected_formatted": f"₹ {total_collected / 100000:.1f}L"
                if total_collected >= 100000
                else f"₹ {total_collected:,.0f}",
                "pending": float(total_pending),
                "pending_formatted": f"₹ {total_pending / 100000:.1f}L"
                if total_pending >= 100000
                else f"₹ {total_pending:,.0f}",
                "pending_count": pending_count,
                "revenue": float(total_revenue),
                "revenue_formatted": f"₹ {total_revenue / 100000:.1f}L"
                if total_revenue >= 100000
                else f"₹ {total_revenue:,.0f}",
            },
            "chart": {"labels": labels, "data": data_points},
            "transactions": transactions,
            "pending_students": pending_list,
            "collected_students": collected_list,
            "revenue_students": revenue_list,
            "course_breakdown": course_breakdown,
            "staff_breakdown": staff_breakdown,
        }
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def student_history(request, student_id):
    history = StudentChangeLog.objects.filter(student_id=student_id).order_by(
        "-changed_at"
    )

    data = []
    for h in history:
        data.append(
            {
                "id": h.id,
                "field_name": h.field_name,
                "old_value": h.old_value,
                "new_value": h.new_value,
                "changed_by": h.changed_by,
                "changed_at": h.changed_at.strftime("%Y-%m-%d %H:%M"),
            }
        )

    return JsonResponse(data, safe=False)


@api_view(["GET"])
@permission_classes([AllowAny])
def students_by_month(request):
    """
    Returns all admitted students sorted by admission date.
    Used for the Student Details 'Students by Month' view.
    """
    students = Student.objects.filter(is_deleted=False).order_by("-admission_date")
    serializer = StudentSerializer(students, many=True)
    return Response(serializer.data)


class UserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        user_profile = user.userprofile
        if user_profile.role == "SUPERADMIN":
            return User.objects.all().order_by("-date_joined")
        return User.objects.all().order_by("-date_joined")

    def create(self, request, *args, **kwargs):
        if request.user.userprofile.role != "SUPERADMIN":
            return Response({"error": "Only SUPERADMIN can create users"}, status=403)

        response = super().create(request, *args, **kwargs)
        if response.status_code == 201:
            new_user_name = response.data.get("username", "N/A")
            # Step 10: LOG USER CREATION
            log_activity(request.user, "USER CREATED", f"Created user {new_user_name}")
        return response

    def destroy(self, request, *args, **kwargs):
        if request.user.userprofile.role != "SUPERADMIN":
            return Response({"error": "Only SUPERADMIN can delete users"}, status=403)

        user_to_delete = self.get_object()
        if user_to_delete.username in ["ADMIN", "SUPERADMIN"]:
            return Response({"error": "Core admin users cannot be deleted"}, status=400)

        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=["put"], url_path="role")
    def change_role(self, request, pk=None):
        if request.user.userprofile.role != "SUPERADMIN":
            return Response({"error": "Only SUPERADMIN can change roles"}, status=403)

        user = self.get_object()
        new_role = request.data.get("role")
        if new_role not in [r[0] for r in UserProfile.ROLE_CHOICES]:
            return Response({"error": "Invalid role"}, status=400)

        profile = user.userprofile
        profile.role = new_role
        profile.save()
        return Response({"message": "Role updated successfully"})


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def logout_view(request):
    # Step 5: LOG LOGOUT
    log_activity(request.user, "LOGOUT", f"{request.user.username} logged out")
    from django.contrib.auth import logout

    logout(request)
    return Response({"message": "Logged out successfully"})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_activity_logs(request):
    user = request.user
    user_profile = user.userprofile

    if user_profile.role not in ["ADMIN", "SUPERADMIN", "AM"]:
        return Response({"detail": "Permission denied."}, status=403)

    if user_profile.role == "SUPERADMIN":
        logs = ActivityLog.objects.all().order_by("-created_at")[:300]
    else:
        logs = ActivityLog.objects.all().order_by("-created_at")[:300]
    data = []
    for log in logs:
        data.append(
            {
                "timestamp": log.created_at.strftime("%Y-%m-%d %H:%M:%S"),
                "user": log.user.username if log.user else "System",
                "action": log.action,
                "description": log.description,
                "status": "Success",
            }
        )
    return Response(data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def student_timeline(request, roll_no):
    """Get all activity logs related to a student by their roll number (student_id)."""
    # Filter logs where the description contains the roll number
    user = request.user
    user_profile = user.userprofile

    logs = ActivityLog.objects.filter(description__icontains=roll_no)
    if user_profile.role != "SUPERADMIN":
        logs = logs.filter(company_id=user_profile.company_id)

    logs = logs.order_by("-created_at")

    data = []
    for log in logs:
        data.append(
            {
                "timestamp": log.created_at.strftime("%d %b %Y %H:%M"),
                "user": log.user.username if log.user else "System",
                "action": log.action,
                "description": log.description,
                "status": "Success",
            }
        )
    return Response(data)


# ========== Step 7: Invoice & Receipt API Endpoints ==========


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_invoice(request, student_id):
    """Get invoice for a student by student DB ID."""
    user = request.user
    user_profile = user.userprofile
    query = Invoice.objects.filter(student_id=student_id)
    if user_profile.role != "SUPERADMIN":
        query = query.filter(company_id=user_profile.company_id)

    invoices = query.order_by("-created_at")
    if not invoices.exists():
        return Response({"error": "No Invoice found"}, status=status.HTTP_404_NOT_FOUND)

    # Return the latest invoice
    invoice = invoices.first()
    return Response(
        {
            "invoice_number": invoice.invoice_number,
            "total_fee": invoice.total_fee,
            "discount": invoice.discount,
            "final_fee": invoice.final_fee,
            "date": invoice.created_at.strftime("%Y-%m-%d %H:%M"),
        }
    )


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def get_student_receipts(request, student_id):
    """Get all receipts for a student by student DB ID, handling old records with auto-creation."""
    user = request.user
    user_profile = user.userprofile
    try:
        # Filter receipts for the student
        query = Receipt.objects.filter(student_id=student_id)
        if user_profile.role != "SUPERADMIN":
            query = query.filter(company_id=user_profile.company_id)
        receipts = query.order_by("-created_at")

        if receipts.exists():
            data = []
            for r in receipts:
                data.append(
                    {
                        "id": r.id,
                        "receipt_no": r.receipt_number,
                        "amount": r.amount_paid,
                        "payment_mode": r.payment_mode,
                        "date": r.created_at.strftime("%Y-%m-%d %H:%M"),
                    }
                )
            return Response({"status": "success", "receipts": data})

        # ----------------------------------
        # AUTO CREATE RECEIPT FOR OLD STUDENTS
        # ----------------------------------
        student = Student.objects.get(id=student_id)

        # Extract total_paid from installments JSON (standard ERP logic)
        total_paid = 0
        if isinstance(student.installments, dict):
            total_paid = student.installments.get("total_paid", 0)

        if total_paid and float(total_paid) > 0:
            # Generate a systematic receipt number based on Student ID
            receipt_no = f"REC{student.student_id}"

            # Ensure uniqueness (if student already had a receipt with this ID for some reason)
            if not Receipt.objects.filter(receipt_number=receipt_no).exists():
                new_receipt = Receipt.objects.create(
                    student=student,
                    receipt_number=receipt_no,
                    amount_paid=int(total_paid),
                    payment_mode=student.payment_mode or "CASH",
                    company_id=user_profile.company_id,
                )

                return Response(
                    {
                        "status": "success",
                        "receipts": [
                            {
                                "id": new_receipt.id,
                                "receipt_no": new_receipt.receipt_number,
                                "amount": new_receipt.amount_paid,
                                "payment_mode": new_receipt.payment_mode,
                                "date": new_receipt.created_at.strftime(
                                    "%Y-%m-%d %H:%M"
                                ),
                            }
                        ],
                    }
                )

        return Response(
            {"status": "empty", "message": "No receipts available for this student."}
        )
    except Exception as e:
        return Response({"status": "error", "message": str(e)}, status=400)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def generate_missing_invoices(request):
    """One-time utility: Create invoices for all existing students that don't have one."""
    user_profile = request.user.userprofile
    if user_profile.role == "SUPERADMIN":
        students = Student.objects.filter(is_deleted=False)
    else:
        students = Student.objects.filter(is_deleted=False)
    created = 0

    for student in students:
        if not Invoice.objects.filter(student=student).exists():
            year = datetime.now().year
            count = Invoice.objects.count() + 1
            invoice_no = f"INV{year}{count:04d}"

            Invoice.objects.create(
                student=student,
                invoice_number=invoice_no,
                total_fee=int(student.total_fee),
                discount=int(student.discount),
                final_fee=int(student.final_fee),
                company_id=student.company_id,
            )
            created += 1

    return Response({"message": "Invoices Generated", "count": created})


@api_view(["GET"])
@permission_classes([AllowAny])
def download_receipt_pdf(request, receipt_id):
    """Download receipt as PDF with professional Erpion ERP layout."""
    try:
        receipt = Receipt.objects.get(id=receipt_id)
        student = receipt.student

        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        # --- HEADER SECTION ---
        p.setFont("Helvetica-Bold", 18)
        p.drawCentredString(width / 2, height - 50, "Erpion ERP")

        p.setFont("Helvetica-Bold", 14)
        p.drawCentredString(width / 2, height - 70, "Admission Fee Payment Receipt")

        # Horizontal Line
        p.line(50, height - 85, width - 50, height - 85)

        # Contact Info
        p.setFont("Helvetica", 10)
        p.drawString(50, height - 105, "Address : Erpion ERP")
        p.drawString(
            50, height - 120, "Phone   : +91 80500 00000 (Sample)"
        )  # Generic placeholder as requested
        p.drawString(width - 250, height - 105, "Email   : info@erpion.com")
        p.drawString(width - 250, height - 120, "Website : www.erpion.com")

        # Horizontal Line
        p.line(50, height - 135, width - 50, height - 135)

        # Receipt Metadata
        p.setFont("Helvetica", 11)
        p.drawString(50, height - 155, f"Receipt No : {receipt.receipt_number}")
        p.drawRightString(
            width - 50,
            height - 155,
            f"Date : {receipt.created_at.strftime('%d-%m-%Y')}",
        )

        # Horizontal Line
        p.line(50, height - 170, width - 50, height - 170)

        # --- STUDENT DETAILS SECTION ---
        y_pos = height - 200
        p.setFont("Helvetica-Bold", 13)
        p.drawString(50, y_pos, "STUDENT DETAILS")
        y_pos -= 5
        p.line(50, y_pos, 180, y_pos)

        p.setFont("Helvetica", 11)
        y_pos -= 25
        p.drawString(50, y_pos, f"Student Name : {student.full_name}")
        y_pos -= 20
        p.drawString(50, y_pos, f"Roll Number  : {student.student_id}")

        # Course info handling
        course_name = "N/A"
        if (
            student.courses
            and isinstance(student.courses, list)
            and len(student.courses) > 0
        ):
            course_name = student.courses[0].get("course", "N/A")

        y_pos -= 20
        p.drawString(50, y_pos, f"Course       : {course_name}")
        y_pos -= 20
        p.drawString(50, y_pos, f"Batch        : {student.batch or '2026'}")

        # Horizontal Line
        y_pos -= 25
        p.line(50, y_pos, width - 50, y_pos)

        # --- PAYMENT DETAILS SECTION ---
        y_pos -= 30
        p.setFont("Helvetica-Bold", 13)
        p.drawString(50, y_pos, "PAYMENT DETAILS")
        y_pos -= 5
        p.line(50, y_pos, 180, y_pos)

        # Calculations
        total_paid_so_far = 0
        if isinstance(student.installments, dict):
            total_paid_so_far = student.installments.get("total_paid", 0)
        pending_balance = float(student.final_fee) - float(total_paid_so_far)

        p.setFont("Helvetica", 11)
        y_pos -= 25
        p.drawString(50, y_pos, f"Total Fee        : ₹ {int(student.final_fee):,}")
        y_pos -= 20
        p.drawString(50, y_pos, f"Amount Paid      : ₹ {int(receipt.amount_paid):,}")
        y_pos -= 20
        p.setFont("Helvetica-Bold", 11)
        p.drawString(50, y_pos, f"Pending Balance  : ₹ {int(pending_balance):,}")

        p.setFont("Helvetica", 11)
        y_pos -= 25
        p.drawString(50, y_pos, f"Payment Mode : {receipt.payment_mode}")

        # Horizontal Line
        y_pos -= 25
        p.line(50, y_pos, width - 50, y_pos)

        # --- DECLARATION SECTION ---
        y_pos -= 40
        p.setFont("Helvetica-Bold", 11)
        p.drawString(50, y_pos, "DECLARATION")
        p.setFont("Helvetica", 10)
        y_pos -= 20
        p.drawString(50, y_pos, "This is a computer generated receipt.")
        y_pos -= 15
        p.drawString(50, y_pos, "No signature required.")

        # --- FOOTER ---
        p.setFont("Helvetica-Bold", 10)
        p.drawCentredString(width / 2, 50, "Generated by Erpion ERP")

        p.showPage()
        p.save()
        buffer.seek(0)

        response = HttpResponse(buffer, content_type="application/pdf")
        filename = f"{student.student_id}.pdf"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response
    except Exception as e:
        return HttpResponse(f"Error generating PDF: {str(e)}", status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def global_search(request):
    """
    Search across Students, Leads, Invoices, Receipts, and Users.
    """
    from django.db.models import Q
    from .models import Student, Lead, Invoice, Receipt
    from django.contrib.auth.models import User

    query = request.GET.get("q", "").strip()
    if not query or len(query) < 2:
        return Response([])

    results = []

    # 1. Search Students
    students_q = Student.objects.filter(
        Q(student_id__icontains=query)
        | Q(full_name__icontains=query)
        | Q(phone__icontains=query)
        | Q(email__icontains=query)
    )
    if request.user.userprofile.role != "SUPERADMIN":
        students_q = students_q.filter(company_id=request.user.userprofile.company_id)
    students = students_q.distinct()[:5]

    for s in students:
        # Extract course name from JSONField
        course_name = ""
        if s.courses and isinstance(s.courses, list) and len(s.courses) > 0:
            first_course = s.courses[0]
            if isinstance(first_course, dict):
                course_name = first_course.get("course", "")
        results.append(
            {
                "type": "student",
                "name": s.full_name,
                "title": s.full_name,
                "subtitle": s.student_id,
                "id": s.id,
                "rollno": s.student_id,
                "phone": s.phone or "",
                "course": course_name,
                "route": "/students",
            }
        )

    # 2. Search Leads
    leads_q = Lead.objects.filter(
        Q(full_name__icontains=query)
        | Q(phone__icontains=query)
        | Q(course_interested__icontains=query)
    )
    if request.user.userprofile.role != "SUPERADMIN":
        leads_q = leads_q.filter(company_id=request.user.userprofile.company_id)
    leads = leads_q.distinct()[:5]

    for l in leads:
        results.append(
            {
                "type": "lead",
                # ...
                "name": l.full_name,
                "title": l.full_name,
                "subtitle": f"{l.phone} • {l.course_interested}",
                "id": l.id,
                "leadId": l.id,
                "route": "/leads/followup",
            }
        )

    # 3. Search Invoices
    invoices_q = Invoice.objects.filter(Q(invoice_number__icontains=query))
    if request.user.userprofile.role != "SUPERADMIN":
        invoices_q = invoices_q.filter(company_id=request.user.userprofile.company_id)
    invoices = invoices_q[:5]

    for inv in invoices:
        results.append(
            {
                "type": "invoice",
                "title": f"Invoice {inv.invoice_number}",
                "name": inv.invoice_number,
                "subtitle": f"Student {inv.student.student_id}",
                "id": inv.student.student_id,
                "rollno": inv.student.student_id,
                "invoiceNo": inv.invoice_number,
                "route": "/students",
                "section": "invoice",
            }
        )

    # 4. Search Receipts
    receipts_q = Receipt.objects.filter(Q(receipt_number__icontains=query))
    if request.user.userprofile.role != "SUPERADMIN":
        receipts_q = receipts_q.filter(company_id=request.user.userprofile.company_id)
    receipts = receipts_q[:5]

    for r in receipts:
        results.append(
            {
                "type": "receipt",
                "title": f"Receipt {r.receipt_number}",
                "name": r.receipt_number,
                "subtitle": f"Student {r.student.student_id}",
                "id": r.student.student_id,
                "rollno": r.student.student_id,
                "receiptNo": r.receipt_number,
                "route": "/students",
                "section": "receipt",
            }
        )

    # 5. Search Users
    users_q = User.objects.filter(Q(username__icontains=query))
    if request.user.userprofile.role != "SUPERADMIN":
        users_q = users_q.filter(
            userprofile__company_id=request.user.userprofile.company_id
        )
    users = users_q[:5]

    for u in users:
        results.append(
            {
                "type": "user",
                "name": u.username,
                "title": u.username,
                "subtitle": u.userprofile.role
                if hasattr(u, "userprofile")
                else "Staff",
                "id": u.id,
                "route": "/users",
            }
        )

    return Response(results[:10])


# --- SYSTEM BACKUP APIs ---

import os
from django.conf import settings
from django.http import FileResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from lead_enrollment.utils import auto_backup, restore_backup_db


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_backup(request):
    try:
        user_profile = request.user.userprofile
        if user_profile.role not in ["ADMIN", "SUPERADMIN", "AM"]:
            return Response({"error": "Permission denied."}, status=403)

        company_label = None
        if user_profile.role != "SUPERADMIN":
            company_id = user_profile.company_id
            try:
                company = Company.objects.get(id=company_id)
                company_label = company.company_name.replace(" ", "_").lower()
            except:
                company_label = f"comp_{company_id}"

        filename = auto_backup(company_label=company_label)
        return Response(
            {"status": "success", "message": f"Backup {filename} created successfully"}
        )
    except Exception as e:
        return Response({"status": "error", "message": str(e)}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def backup_list(request):
    user_profile = request.user.userprofile
    if user_profile.role not in ["ADMIN", "SUPERADMIN", "AM"]:
        return Response({"error": "Permission denied."}, status=403)

    backups_dir = os.path.join(settings.BASE_DIR, "backups")
    if not os.path.exists(backups_dir):
        return Response([])

    files = [f for f in os.listdir(backups_dir) if f.endswith(".sql")]

    if user_profile.role != "SUPERADMIN":
        company_id = user_profile.company_id
        try:
            company = Company.objects.get(id=company_id)
            company_label = company.company_name.replace(" ", "_").lower()
        except:
            company_label = f"comp_{company_id}"

        prefix = f"backup_company_{company_label}"
        files = [f for f in files if f.startswith(prefix)]

    # Sort files by name, descending (most recent first)
    files.sort(reverse=True)

    backup_data = []
    for f in files:
        filepath = os.path.join(backups_dir, f)
        size = os.path.getsize(filepath)
        backup_data.append(
            {"filename": f, "size": size, "created_at": os.path.getmtime(filepath)}
        )

    return Response(backup_data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def download_backup(request, filename):
    user_profile = request.user.userprofile
    if user_profile.role not in ["ADMIN", "SUPERADMIN", "AM"]:
        return Response({"error": "Permission denied."}, status=403)

    backups_dir = os.path.join(settings.BASE_DIR, "backups")
    filepath = os.path.join(backups_dir, filename)

    if os.path.exists(filepath):
        response = FileResponse(
            open(filepath, "rb"), as_attachment=True, filename=filename
        )
        return response
    return Response({"error": "File not found"}, status=404)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def restore_backup_api(request):
    if request.user.userprofile.role != "SUPERADMIN":
        return Response({"success": False, "message": "Access denied"})

    file_name = request.POST.get("file")
    if not file_name:
        return Response({"success": False, "message": "No file specified"})

    success = restore_backup_db(file_name)

    if success:
        return Response({"success": True, "message": "Database restored successfully"})

    return Response({"success": False, "message": "Restore failed"})


@api_view(["DELETE"])
@permission_classes([IsAuthenticated])
def delete_backup(request, filename):
    user_profile = request.user.userprofile
    if user_profile.role not in ["ADMIN", "SUPERADMIN", "AM"]:
        return Response({"error": "Permission denied."}, status=403)

    backups_dir = os.path.join(settings.BASE_DIR, "backups")
    filepath = os.path.join(backups_dir, filename)

    if os.path.exists(filepath):
        try:
            os.remove(filepath)
            return Response(
                {"status": "success", "message": f"Backup {filename} deleted"}
            )
        except Exception as e:
            return Response({"status": "error", "message": str(e)}, status=500)
    return Response({"error": "File not found"}, status=404)


class StaffUserViewSet(viewsets.ModelViewSet):
    serializer_class = StaffUserSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        from lead_enrollment.models import UserProfile

        db = getattr(self.request, "tenant_db", "default")
        if not db or db == "default":
            return StaffUser.objects.none()

        # MANUALLY check permission in tenant DB to avoid erpion_main crash
        profile = (
            UserProfile.objects
            .filter(user__email=self.request.user.email)
            .first()
        )
        if not profile or profile.role != "ADMIN":
            return StaffUser.objects.none()

        return (
            StaffUser.objects
            .select_related("user")
            .all()
            .order_by("-created_at")
        )

    def list(self, request, *args, **kwargs):
        from lead_enrollment.models import UserProfile

        try:
            db = getattr(request, "tenant_db", "default")
            if not db or db == "default":
                return Response({"error": "Tenant database not assigned"}, status=400)

            # MANUALLY check permission in tenant DB
            profile = (
                UserProfile.objects
                .filter(user__email=request.user.email)
                .first()
            )
            if not profile or profile.role != "ADMIN":
                return Response({"error": "Access Denied"}, status=403)

            users = (
                StaffUser.objects
                .select_related("user")
                .all()
                .order_by("-created_at")
            )

            # Manual construction as per user pattern (ensuring email comes from linked user)
            data = []
            for u in users:
                data.append(
                    {
                        "id": u.id,
                        "full_name": u.full_name,
                        "email": u.user.email,  # Email comes from the linked Django User
                        "phone": u.phone,
                        "role": u.role,
                        "status": u.status,
                        "created_at": u.created_at,
                    }
                )

            return Response(data)

        except Exception as e:
            return Response({"error": str(e)}, status=500)

    def create(self, request, *args, **kwargs):
        from django.contrib.auth.models import User
        from django.db import transaction
        from saas_management.models import GlobalUser, Company, CompanyUserMapping

        try:
            db = getattr(request, "tenant_db", "default")
            if not db or db == "default":
                return Response({"error": "Tenant database not assigned"}, status=400)

            # MANUALLY check permission in tenant DB to avoid erpion_main crash
            from lead_enrollment.models import UserProfile

            profile = (
                UserProfile.objects
                .filter(user__email=request.user.email)
                .first()
            )
            if not profile or profile.role != "ADMIN":
                return Response(
                    {"error": "Only Company Admins can manage staff."}, status=403
                )

            full_name = request.data.get("full_name")
            email = request.data.get("email")
            phone = request.data.get("phone")
            password = request.data.get("password")
            role = request.data.get("role")
            status_value = request.data.get("status", "Active")

            # 3. FIX DUPLICATE EMAIL ERROR
            # Check in tenant DB
            if User.objects.filter(username=email).exists():
                return Response(
                    {
                        "error": "Staff member with this email already exists in this company."
                    },
                    status=400,
                )

            # Check in Global DB
            if GlobalUser.objects.filter(email=email).exists():
                # If they already exist globally but not in this tenant, we map them (allowing cross-tenant users? User didn't specify, but for now strict isolation)
                return Response(
                    {
                        "error": "A user with this email is already registered in the system."
                    },
                    status=400,
                )

            with transaction.atomic(using=db):
                # Step 1 — Create Django User in Tenant DB
                user = User.objects.db_manager(db).create_user(
                    username=email, email=email, password=password
                )

                # Step 2 — Create Staff Profile (StaffUser)
                staff_user = StaffUser.objects.create(
                    user=user,
                    full_name=full_name,
                    phone=phone,
                    role=role,
                    status=status_value,
                )

                # Step 3 — Sync to Global SaaS Auth (erpion_main)
                company = (
                    Company.objects
                    .filter(database_name=db)
                    .first()
                )
                if company:
                    # Sync GlobalUser
                    GlobalUser.objects.update_or_create(
                        username=email,
                        defaults={
                            "email": email,
                            "password": user.password,  # Use the hashed password from the user object
                            "role": "STAFF",
                        },
                    )

                    # Sync Company Mapping
                    CompanyUserMapping.objects.update_or_create(
                        user_email=email, defaults={"company": company, "role": role}
                    )

                    # Create AuthUser in erpion_main so BearerTokenAuthentication finds it
                    global_auth_user, created = User.objects.using(
                        "erpion_main"
                    ).get_or_create(username=email, defaults={"email": email})
                    global_auth_user.password = user.password
                    global_auth_user.save(using="erpion_main")

            return Response({"message": "User created successfully"}, status=201)

        except Exception as e:
            import traceback

            traceback.print_exc()
            return Response({"error": str(e)}, status=500)

    def update(self, request, *args, **kwargs):
        from django.contrib.auth.models import User
        from django.db import transaction
        from saas_management.models import GlobalUser, CompanyUserMapping

        try:
            db = getattr(request, "tenant_db", "default")
            instance = self.get_object()  # Fetching StaffUser
            user = instance.user  # Accessing the linked auth_user

            full_name = request.data.get("full_name", instance.full_name)
            phone = request.data.get("phone", instance.phone)
            role = request.data.get("role", instance.role)
            status_value = request.data.get("status", instance.status)

            password = request.data.get("password")

            with transaction.atomic(using=db):
                # Update auth_user in tenant DB if password changed
                if password:
                    user.set_password(password)
                    user.save(using=db)

                # Update staff profile
                instance.full_name = full_name
                instance.phone = phone
                instance.role = role
                instance.status = status_value
                instance.save(using=db)

                # Sync updates to erpion_main
                GlobalUser.objects.filter(email=user.email).update(
                    password=user.password, role="STAFF"
                )
                User.objects.filter(email=user.email).update(
                    password=user.password
                )
                CompanyUserMapping.objects.filter(
                    user_email=user.email
                ).update(role=role)

            return Response({"message": "User updated successfully"})
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    def destroy(self, request, *args, **kwargs):
        from django.contrib.auth.models import User
        from saas_management.models import GlobalUser, CompanyUserMapping

        try:
            db = getattr(request, "tenant_db", "default")
            instance = self.get_object()

            user_id = instance.user_id
            email = instance.user.email

            # Step 1: Delete from staff_users first to resolve FK constraint
            instance.delete(using=db)

            # Step 2: Delete from auth_user
            User.objects.filter(id=user_id).delete()

            # Sync deletion from SaaS main DB
            GlobalUser.objects.filter(email=email).delete()
            CompanyUserMapping.objects.filter(
                user_email=email
            ).delete()
            User.objects.filter(email=email).delete()

            return Response({"message": "User deleted successfully"})
        except Exception as e:
            return Response({"error": f"Failed to delete user: {str(e)}"}, status=500)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def tenant_global_search(request):
    """
    New API for tenant-specific global search across Students, Leads, and Staff.
    Supports suggestions for Roll No, Name, Phone, and Email.
    Automatically routes to the correct tenant database via middleware.
    """
    from django.db.models import Q
    from .models import Student, Lead, StaffUser

    query = request.query_params.get("q", "").strip()
    if not query or len(query) < 2:
        return Response([])

    results = []

    # 1. Search Students (in current tenant DB)
    students = Student.objects.filter(
        Q(student_id__icontains=query)
        | Q(full_name__icontains=query)
        | Q(phone__icontains=query)
        | Q(email__icontains=query),
        is_deleted=False,
    ).distinct()[:5]

    for s in students:
        results.append(
            {
                "type": "student",
                "name": s.full_name,
                "roll_no": s.student_id,
                "phone": s.phone,
                "id": s.id,
            }
        )

    # 2. Search Leads
    leads = Lead.objects.filter(
        Q(full_name__icontains=query)
        | Q(phone__icontains=query)
        | Q(email__icontains=query)
    ).distinct()[:5]

    for l in leads:
        results.append(
            {"type": "lead", "name": l.full_name, "phone": l.phone, "id": l.id}
        )

    # 3. Search Staff
    # Note: StaffUser links to auth.User for email/name parts
    staff = (
        StaffUser.objects.filter(
            Q(full_name__icontains=query)
            | Q(user__email__icontains=query)
            | Q(user__first_name__icontains=query)
            | Q(user__last_name__icontains=query)
        )
        .select_related("user")
        .distinct()[:5]
    )

    for st in staff:
        results.append(
            {"type": "staff", "name": st.full_name, "email": st.user.email, "id": st.id}
        )

    return Response(results)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def today_followups(request):
    """
    Returns today's pending follow-ups for the logged-in staff.
    """
    from datetime import date
    from django.db.models import Q
    from .models import Lead

    today = date.today()
    # Filter for today, assigned to current user, status is 'Follow-up' (pending)
    leads = Lead.objects.filter(next_follow_up_date=today, status="Follow-up").order_by(
        "next_follow_up_time"
    )[:20]

    # Non-admin users see only their assigned leads
    if request.user.status != "SUPERADMIN" and request.user.userprofile.role not in [
        "ADMIN",
        "AM",
    ]:
        leads = leads.filter(assigned_to=request.user.username)

    results = []
    for l in leads:
        results.append(
            {
                "lead_id": l.id,
                "lead_name": l.full_name,
                "course": l.course_interested,
                "phone": l.phone,
                "time": l.next_follow_up_time.strftime("%H:%M")
                if l.next_follow_up_time
                else "No Time",
            }
        )
    return Response(results)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def pending_followup_count(request):
    """
    Returns number of pending follow-ups for the day.
    """
    from datetime import date
    from .models import Lead

    today = date.today()
    leads_q = Lead.objects.filter(next_follow_up_date=today, status="Follow-up")

    if request.user.status != "SUPERADMIN" and request.user.userprofile.role not in [
        "ADMIN",
        "AM",
    ]:
        leads_q = leads_q.filter(assigned_to=request.user.username)

    count = leads_q.count()
    return Response({"count": count})


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def overdue_followups(request):
    """
    Returns follow-ups where time is past current time and status is pending.
    """
    from datetime import datetime
    from .models import Lead

    now = datetime.now()
    today = now.date()
    current_time = now.time()

    # Overdue = (Date is today and Time is past) OR Date is in the past
    # And status is 'Follow-up'
    leads = (
        Lead.objects.filter(status="Follow-up")
        .filter(
            (Q(next_follow_up_date=today) & Q(next_follow_up_time__lt=current_time))
            | Q(next_follow_up_date__lt=today)
        )
        .order_by("next_follow_up_date", "next_follow_up_time")[:20]
    )

    if request.user.status != "SUPERADMIN" and request.user.userprofile.role not in [
        "ADMIN",
        "AM",
    ]:
        leads = leads.filter(assigned_to=request.user.username)

    results = []
    for l in leads:
        results.append(
            {
                "lead_id": l.id,
                "lead_name": l.full_name,
                "course": l.course_interested,
                "time": f"{l.next_follow_up_date.strftime('%Y-%m-%d')} {l.next_follow_up_time.strftime('%H:%M') if l.next_follow_up_time else '00:00'}",
                "status": "OVERDUE",
            }
        )
    return Response(results)


@api_view(["POST"])
@permission_classes([AllowAny])
def student_login(request):
    """
    Login for students using Portal Code, Email and Date of Birth (DOB).
    """
    from django.conf import settings
    import jwt
    from datetime import datetime, timedelta
    from backend.middleware import _ensure_db_registered
    from saas_management.models import Company

    portal_code = request.data.get("portal_code")
    email = request.data.get("email")
    dob = request.data.get("dob")  # Format: YYYY-MM-DD

    if not portal_code or not email or not dob:
        return Response(
            {"error": "Institute Code, Email and DOB are required"},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Resolve database name
    tenant_db = f"erpion_tenant_{portal_code}"

    try:
        # Check if institute exists in erpion_main (Privacy check)
        institute_exists = (
            Company.objects
            .filter(database_name=tenant_db)
            .exists()
        )
        if not institute_exists:
            return Response(
                {"error": "Invalid institute portal code"},
                status=status.HTTP_404_NOT_FOUND,
            )

        # 1. Register tenant database if needed
        _ensure_db_registered(tenant_db)

        # 2. Check if student exists in the selected tenant database
        student = (
            Student.objects
            .filter(email__iexact=email, dob=dob)
            .first()
        )
        if not student:
            return Response(
                {"error": "Invalid login credentials"},
                status=status.HTTP_401_UNAUTHORIZED,
            )

        # Generate Token
        payload = {
            "student_id": student.id,
            "roll_no": student.student_id,
            "email": student.email,
            "role": "STUDENT",
            "tenant_db": tenant_db,
            "exp": datetime.utcnow() + timedelta(days=2),
        }
        token = jwt.encode(payload, settings.SECRET_KEY, algorithm="HS256")

        return Response(
            {
                "token": token,
                "role": "STUDENT",
                "tenant_db": tenant_db,
                "student": {
                    "id": student.id,
                    "name": student.full_name,
                    "roll_no": student.student_id,
                    "email": student.email,
                },
            }
        )
    except Exception as e:
        return Response(
            {"error": "Connection error or internal server error."}, status=500
        )


def get_student_from_token(request):
    """Helper to get student from JWT token."""
    import jwt
    from django.conf import settings

    auth_header = request.headers.get("Authorization")
    if not auth_header or not auth_header.startswith("Bearer "):
        return None
    token = auth_header.split(" ")[1]
    try:
        payload = jwt.decode(token, settings.SECRET_KEY, algorithms=["HS256"])
        if payload.get("role") != "STUDENT":
            return None
        db = payload.get("tenant_db")
        student_id = payload.get("student_id")
        return Student.objects.filter(id=student_id).first(), db
    except:
        return None


@api_view(["GET"])
@permission_classes([AllowAny])
def student_portal_dashboard(request):
    from django.db.models import Sum

    res = get_student_from_token(request)
    if not res:
        return Response({"error": "Unauthorized"}, status=401)
    student, db = res

    # Courses
    course_info = student.courses[0] if student.courses else {}
    course_name = course_info.get("course", "N/A")
    batch = course_info.get("batch_timing", "N/A")
    trainer = course_info.get("trainer", "N/A")

    return Response(
        {
            "name": student.full_name,
            "roll_no": student.student_id,
            "course": course_name,
            "batch": batch,
            "trainer": trainer,
            "fee_status": student.fee_status,
            "certificate_status": student.certificate_status,
            "final_fee": student.final_fee,
            "total_paid": Receipt.objects
            .filter(student=student)
            .aggregate(Sum("amount_paid"))["amount_paid__sum"]
            or 0,
        }
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def student_portal_payments(request):
    res = get_student_from_token(request)
    if not res:
        return Response({"error": "Unauthorized"}, status=401)
    student, db = res

    receipts = Receipt.objects.filter(student=student).order_by("-created_at")
    receipt_data = [
        {
            "receipt_number": r.receipt_number,
            "amount": r.amount_paid,
            "mode": r.payment_mode,
            "date": r.created_at.strftime("%Y-%m-%d"),
        }
        for r in receipts
    ]

    total_paid = sum(r["amount"] for r in receipt_data)

    return Response(
        {
            "total_fee": student.total_fee,
            "discount": student.discount,
            "final_fee": student.final_fee,
            "paid_amount": total_paid,
            "pending_amount": student.final_fee - total_paid,
            "history": receipt_data,
        }
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def student_portal_profile(request):
    res = get_student_from_token(request)
    if not res:
        return Response({"error": "Unauthorized"}, status=401)
    student, db = res
    return Response(
        {
            "name": student.full_name,
            "roll_no": student.student_id,
            "email": student.email,
            "phone": student.phone,
            "address": student.address,
            "dob": student.dob,
        }
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def student_portal_course(request):
    res = get_student_from_token(request)
    if not res:
        return Response({"error": "Unauthorized"}, status=401)
    student, db = res

    # Assuming first course for now
    course_info = student.courses[0] if student.courses else {}

    return Response(
        {
            "course_name": course_info.get("course", "N/A"),
            "batch_timing": course_info.get("batch_timing", "Morning (9 AM - 11 AM)"),
            "duration": course_info.get("duration", "3 Months"),
            "trainer": course_info.get("trainer", "Senior Technical Lead"),
        }
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def student_portal_certificate(request):
    res = get_student_from_token(request)
    if not res:
        return Response({"error": "Unauthorized"}, status=401)
    student, db = res

    # Check if certificate exists (using a placeholder logic for now)
    # The model has certificate_applied and certificate_status
    has_certificate = (
        student.certificate_status == "Graduated" or student.certificate_applied
    )

    return Response(
        {
            "status": student.certificate_status,
            "has_certificate": has_certificate,
            "download_url": f"/api/student/download-certificate/{student.id}/"
            if has_certificate
            else None,
        }
    )


@api_view(["GET"])
@permission_classes([AllowAny])
def public_institutes_list(request):
    """
    Public API to fetch list of ACTIVE institutes (companies) from erpion_main.
    """
    from saas_management.models import Company

    try:
        companies = (
            Company.objects
            .filter(status="ACTIVE")
            .values("id", "company_name", "database_name")
        )
        data = []
        for c in companies:
            data.append(
                {
                    "id": c["id"],
                    "name": c["company_name"],
                    "tenant_db": c["database_name"],
                }
            )
        return Response(data)
    except Exception as e:
        return Response({"error": str(e)}, status=500)


from openpyxl import Workbook
from django.http import HttpResponse
from django.db.models import Count


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_leads_excel(request):
    """
    Exports all leads from the current tenant database into an Excel (.xlsx) file.
    Includes FULL historical follow-up remarks for every lead.
    """
    try:
        from openpyxl.styles import Alignment, Font

        # Get active tenant database from current request context
        leads = Lead.objects.annotate(walkin_count=Count("walk_ins")).order_by(
            "-created_at"
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "Leads"

        # Headers
        headers = [
            "Lead Name",
            "Course",
            "Phone Number",
            "Lead Status",
            "Walk-in Count",
            "Next Follow-up Date",
            "Follow-up Remarks (Full History)",
            "Assigned Staff",
            "Lead Source",
            "Created Date",
        ]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True)

        for lead in leads:
            # 1. Fetch all remarks from LeadRemark history
            history = lead.remarks_history.all().order_by("created_at")

            # 2. Build the full remarks string
            remarks_list = []

            # Include initial remark if exists
            if lead.remarks:
                initial_date = lead.created_at.strftime("%Y-%m-%d")
                remarks_list.append(f"{initial_date} - (Initial) {lead.remarks}")

            # Include historical remarks
            for r in history:
                r_date = r.created_at.strftime("%Y-%m-%d")
                remarks_list.append(f"{r_date} - {r.remark_text}")

            full_remarks_text = (
                "\n".join(remarks_list) if remarks_list else "No remarks"
            )

            ws.append(
                [
                    lead.full_name,
                    lead.course_interested,
                    lead.phone,
                    lead.status,
                    getattr(lead, "walkin_count", 0),
                    lead.next_follow_up_date.strftime("%Y-%m-%d")
                    if lead.next_follow_up_date
                    else "N/A",
                    full_remarks_text,
                    lead.assigned_to if lead.assigned_to else "Unassigned",
                    lead.source if lead.source else "Direct",
                    lead.created_at.strftime("%Y-%m-%d %H:%M"),
                ]
            )

        # Formatting: Enable wrap text for the Remarks column (Column G, index 7)
        for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Adjust column widths
        for column_cells in ws.columns:
            col_letter = column_cells[0].column_letter
            if col_letter == "G":  # Remarks column
                ws.column_dimensions[col_letter].width = 60
            else:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[col_letter].width = min(
                    length + 2, 40
                )  # Cap width for others

        # Dynamic Filename Logic
        from datetime import datetime
        from saas_management.models import Company

        current_date = datetime.now().strftime("%Y-%m-%d")
        db_name = getattr(request, "tenant_db", "default")
        company_prefix = "erpion"

        if db_name.startswith("erpion_tenant_"):
            try:
                company = Company.objects.get(
                    database_name=db_name
                )
                company_prefix = company.company_name.lower().replace(" ", "_")
            except Exception:
                # Fallback to the portal code if DB not found in erpion_main
                company_prefix = db_name.replace("erpion_tenant_", "")
        elif db_name == "erpion_main":
            company_prefix = "saas_admin"

        filename = f"{company_prefix}_leads_{current_date}.xlsx"

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)
        return response

    except Exception as e:
        print(f"DEBUG: Lead Export Failed: {str(e)}")
        return HttpResponse(f"Export Error: {str(e)}", status=500)
